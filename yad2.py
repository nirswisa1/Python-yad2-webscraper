from selenium import webdriver
import smtplib
from email.message import EmailMessage
from openpyxl import load_workbook
from tkinter import *
#Web scraping for "yad2" website

#tkinter - GUI
root = Tk()
root.title('Nir Yad2 Bot')
url = ''
email = ''


def myClick():
    global url
    global email
    url = urlE.get()
    email = emailE.get()
    root.destroy()
    return url, email


# slots
urlE = Entry(root, width=50, borderwidth=2)
urlE.grid(row=0, column=1)
emailE = Entry(root, width=50, borderwidth=2)
emailE.grid(row=1, column=1)
# labels
label1 = Label(root, text='Url: ')
label2 = Label(root, text='email: ')
label1.grid(row=0, column=0)
label2.grid(row=1, column=0)
# Button
b = Button(root, text="Run", command=myClick)
b.grid(row=2, column=2)

root.mainloop()
#Selenium
driver = webdriver.Chrome(r'C:\Users\NS\Downloads\PYTHON\chromedriver.exe')

driver.get(str(url))

driver.maximize_window()
driver.implicitly_wait(300)

ids = []
#Loading excel data base
wb = load_workbook('yad.xlsx')
ws = wb.active
#Loading existing apts
L = [i.value for i in ws['A'] if i.value]


def my_app_bot(ids_of_apt):
    global L
    global ids
    global ws
    global wb

    driver.refresh()
    new_apts = []
    scrolling = 400
    for i in range(0, 20):
        scrolling += 400
        print(i)
        iStr = str(i)

        if driver.find_element_by_id('feed_item_' + iStr + '_price'):
            apt_id = driver.find_element_by_id('feed_item_' + iStr + '_price')
            apt_id.click()
            ad_num = driver.find_element_by_class_name('num_ad').text[-8:]
            driver.execute_script("window.scrollTo(0, %s)" % str(scrolling))

            if ad_num not in ids_of_apt:
                apt_link = driver.find_element_by_xpath(
                    """//*[@id="accordion_wide_%s"]/div/div[3]/ul/li[4]/a""" % iStr).get_attribute('href')
                new_apts.append(apt_link)
                ws["A"+str(ws.max_row+1)] = ad_num
                wb.save('yad.xlsx')

        else:
            print('no')
        apt_id.click()

    if new_apts:
        #send email
        sender = "nirswisa1@gmail.com"
        rec = email
        password = 'pemvbavnbbyurbfx'
        mes = '''
        '''.join(new_apts)
        msg = EmailMessage()
        msg.set_content(mes)

        msg['Subject'] = 'New Apartments!'
        msg['From'] = sender
        msg['To'] = rec

        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(sender, password)
        server.send_message(msg)
        server.quit()
        print('email sent')
    my_app_bot(L)


my_app_bot(L)

